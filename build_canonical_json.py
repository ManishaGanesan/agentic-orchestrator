#!/usr/bin/env python3
"""
build_canonical_json.py

Goal:
  Convert:
    1) Master Excel (User Story title, description, PR link)
    2) Local SQL scripts (downloaded from PR)
    3) Business Excel sheets with highlighted changes
  into a canonical JSON dataset.

Requires:
  pip install pandas openpyxl

Usage:
  python build_canonical_json.py \
    --master "master.xlsx" \
    --sql-root "./downloaded_sql" \
    --business "biz1.xlsx" --business "biz2.xlsx" \
    --out "canonical.json"

Notes:
  - Script attempts to auto-detect key columns in master Excel.
  - SQL scripts are matched by Story ID (e.g., US12345678) and/or PR number in filename/path.
  - Business Excel highlights are detected via cell fill (rgb/theme/indexed).
"""

import argparse
import datetime as dt
import hashlib
import json
import os
from platform import release
import re
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
from openpyxl import load_workbook

# ---------------------------
# Helpers
# ---------------------------

def now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def sha256_text(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8", errors="ignore")).hexdigest()

def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def safe_read_text(path: str) -> str:
    # Many .sql files are UTF-8; some may be cp1252. Try a couple.
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return f.read()
        except Exception:
            pass
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def norm_col(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip().lower())

def first_nonempty(*vals):
    for v in vals:
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            return s
    return ""


def extract_story_and_release(title: str):
    if not title:
        return "", ""

    # Extract story ID (US12345678)
    story_match = re.search(r"(US\s*[-_ ]?\d{5,})", title, re.IGNORECASE)
    story_id = ""
    if story_match:
        story_id = story_match.group(1).replace(" ", "").replace("-", "").upper()

    # Extract release (V20, V21...)
    release_match = re.search(r"(V\d{1,3})", title, re.IGNORECASE)
    release = ""
    if release_match:
        release = release_match.group(1).upper()

    return story_id, release

def extract_pr_number(pr_url: str) -> str:
    """
    Extract PR number from common URL patterns:
      - .../pull/1234
      - .../pullrequest/1234
      - ...pullRequestId=1234
    """
    if not pr_url:
        return ""
    patterns = [
        r"/pull/(\d+)\b",
        r"/pullrequest/(\d+)\b",
        r"pullRequestId=(\d+)\b",
        r"pr[=/](\d+)\b",
    ]
    for p in patterns:
        m = re.search(p, pr_url, flags=re.IGNORECASE)
        if m:
            return m.group(1)
    return ""

def best_col(df_cols: List[str], synonyms: List[str]) -> Optional[str]:
    """
    Find best matching column in df_cols given synonyms.
    Matching is done on normalized names and partial containment.
    """
    normed = {c: norm_col(c) for c in df_cols}
    syn_norm = [norm_col(s) for s in synonyms]

    # exact match
    for c, nc in normed.items():
        if nc in syn_norm:
            return c

    # containment match
    for c, nc in normed.items():
        for s in syn_norm:
            if s in nc or nc in s:
                return c

    return None

# ---------------------------
# SQL parsing (lightweight)
# ---------------------------

SQL_OBJECT_PATTERNS = [
    ("table", r"\b(?:create|alter)\s+table\s+([a-zA-Z0-9_\[\]\.]+)"),
    ("view", r"\b(?:create|alter)\s+view\s+([a-zA-Z0-9_\[\]\.]+)"),
    ("procedure", r"\b(?:create|alter)\s+proc(?:edure)?\s+([a-zA-Z0-9_\[\]\.]+)"),
    ("function", r"\b(?:create|alter)\s+function\s+([a-zA-Z0-9_\[\]\.]+)"),
    ("index", r"\bcreate\s+(?:unique\s+)?index\s+([a-zA-Z0-9_\[\]\.]+)\s+on\s+([a-zA-Z0-9_\[\]\.]+)"),
]

SQL_OP_PATTERNS = [
    ("insert", r"\binsert\s+into\b"),
    ("update", r"\bupdate\b"),
    ("delete", r"\bdelete\b"),
    ("merge", r"\bmerge\b"),
    ("truncate", r"\btruncate\s+table\b"),
]

def parse_sql_summary(sql_text: str) -> Dict[str, Any]:
    t = sql_text or ""
    low = t.lower()

    objects = []
    for kind, pat in SQL_OBJECT_PATTERNS:
        for m in re.finditer(pat, t, flags=re.IGNORECASE):
            if kind == "index":
                idx_name = m.group(1)
                on_tbl = m.group(2)
                objects.append({"type": "index", "name": idx_name, "on": on_tbl})
            else:
                objects.append({"type": kind, "name": m.group(1)})

    operations = []
    for op, pat in SQL_OP_PATTERNS:
        if re.search(pat, low, flags=re.IGNORECASE):
            operations.append(op)

    # crude impacted table references: FROM/JOIN/INTO/UPDATE
    impacted = set()
    for pat in [
        r"\bfrom\s+([a-zA-Z0-9_\[\]\.]+)",
        r"\bjoin\s+([a-zA-Z0-9_\[\]\.]+)",
        r"\binto\s+([a-zA-Z0-9_\[\]\.]+)",
        r"\bupdate\s+([a-zA-Z0-9_\[\]\.]+)",
        r"\bdelete\s+from\s+([a-zA-Z0-9_\[\]\.]+)",
    ]:
        for m in re.finditer(pat, t, flags=re.IGNORECASE):
            impacted.add(m.group(1))

    return {
        "operations": sorted(set(operations)),
        "objects_declared": objects,
        "tables_referenced": sorted(impacted),
        "text_sha256": sha256_text(t),
    }

# ---------------------------
# Finding SQL scripts per story
# ---------------------------

def find_sql_files(sql_root: str, keys: List[str]) -> List[str]:
    """
    Search for .sql files under sql_root whose path/filename contains any key.
    keys should be strings like ["US12345678", "1234"] (PR #), etc.
    """
    if not sql_root or not os.path.isdir(sql_root):
        return []

    keys = [k for k in keys if k]
    if not keys:
        return []

    matches = []
    for root, _, files in os.walk(sql_root):
        for fn in files:
            if not fn.lower().endswith(".sql"):
                continue
            full = os.path.join(root, fn)
            hay = (full + " " + fn).lower()
            if any(k.lower() in hay for k in keys if k):
                matches.append(full)

    # deterministic ordering
    return sorted(set(matches))

# ---------------------------
# Business Excel highlight extraction
# ---------------------------

def is_highlighted(cell) -> Tuple[bool, str]:
    """
    Detect highlight-ish formatting based on fill.
    Returns (is_highlighted, color_descriptor).
    Works for rgb/theme/indexed.
    """
    fill = cell.fill
    if fill is None:
        return (False, "")

    # Many cells will have fill.patternType=None or "none"
    if not getattr(fill, "patternType", None):
        return (False, "")

    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return (False, "")

    # fgColor can be rgb, theme, indexed
    ctype = getattr(fg, "type", None)
    rgb = getattr(fg, "rgb", None)
    theme = getattr(fg, "theme", None)
    indexed = getattr(fg, "indexed", None)

    # treat non-empty non-white RGB as highlighted
    if ctype == "rgb" and rgb:
        r = rgb.upper()
        # common defaults: 00000000 (transparent), FFFFFFFF (white)
        if r not in ("00000000", "FFFFFFFF", "FFFFFF"):
            return (True, f"rgb:{r}")
        return (False, f"rgb:{r}")

    # theme-based fill is often used for highlights; consider as highlighted
    if ctype == "theme" and theme is not None:
        return (True, f"theme:{theme}")

    # indexed colors can also indicate highlight
    if ctype == "indexed" and indexed is not None:
        # indexed 64 is often "no color"
        if int(indexed) != 64:
            return (True, f"indexed:{indexed}")
        return (False, f"indexed:{indexed}")

    # fallback: if patternType exists but unknown color details, treat as highlighted
    return (True, f"type:{ctype or 'unknown'}")

def extract_highlights_from_workbook(
    xlsx_path: str,
    max_rows: int = 5000,
    max_cols: int = 200
) -> Dict[str, Any]:
    """
    Returns a structured dict:
      {
        "workbook": "...",
        "sheets": [
          {
            "sheet": "Sheet1",
            "highlights": [
              {
                "cell": "D12",
                "row": 12,
                "col": 4,
                "header": "Rate",
                "value": "...",
                "color": "rgb:FFFF00",
                "row_snapshot": { "ColAHeader": "...", ... }
              }, ...
            ]
          }, ...
        ]
      }
    """
    wb = load_workbook(filename=xlsx_path, data_only=True)
    out = {"workbook": os.path.basename(xlsx_path), "path": xlsx_path, "sheets": []}

    for ws in wb.worksheets:
        # Determine headers from first row (row=1)
        headers = {}
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(row=1, column=c).value
            if v is None or str(v).strip() == "":
                headers[c] = f"COL_{c}"
            else:
                headers[c] = str(v).strip()

        sheet_highlights = []
        # Scan cells for highlight fill
        for r in range(2, min(ws.max_row, max_rows) + 1):
            # Prepare a row snapshot keyed by headers (so you can reconstruct "what changed")
            row_snapshot = {}
            for c in range(1, min(ws.max_column, max_cols) + 1):
                hv = ws.cell(row=r, column=c).value
                row_snapshot[headers[c]] = hv

            for c in range(1, min(ws.max_column, max_cols) + 1):
                cell = ws.cell(row=r, column=c)
                hl, color = is_highlighted(cell)
                if not hl:
                    continue

                sheet_highlights.append({
                    "cell": cell.coordinate,
                    "row": r,
                    "col": c,
                    "header": headers.get(c, f"COL_{c}"),
                    "value": cell.value,
                    "color": color,
                    "row_snapshot": row_snapshot
                })

        # Only keep sheets that had at least one highlight
        if sheet_highlights:
            out["sheets"].append({
                "sheet": ws.title,
                "highlights_count": len(sheet_highlights),
                "highlights": sheet_highlights
            })

    return out

# ---------------------------
# Master Excel ingestion
# ---------------------------

def read_master_excel(master_path: str, sheet: Optional[str] = None) -> pd.DataFrame:
    if sheet:
        return pd.read_excel(master_path, sheet_name=sheet, engine="openpyxl")
    # default first sheet
    return pd.read_excel(master_path, engine="openpyxl")

def master_rows_to_stories(df: pd.DataFrame) -> List[Dict[str, Any]]:
    cols = list(df.columns)

    col_title = best_col(cols, ["title", "user story title", "story title", "name"])
    col_desc  = best_col(cols, ["description", "user story description", "details", "story description"])
    col_pr    = best_col(cols, ["pr", "pr link", "pull request", "pull request link", "link"])
    col_id    = best_col(cols, ["formatted id", "id", "story id", "user story id", "us id", "user story"])

    stories = []
    for _, row in df.iterrows():
        title = first_nonempty(row.get(col_title) if col_title else None)
        desc  = first_nonempty(row.get(col_desc) if col_desc else None)
        prurl = first_nonempty(row.get(col_pr) if col_pr else None)
        given_id = first_nonempty(row.get(col_id) if col_id else None)

        story_id, release = extract_story_and_release(title)
        pr_num = extract_pr_number(prurl)

        if not any([title, desc, prurl, story_id]):
            continue
  
        stories.append({
            "story_id": story_id,
            "release": release,
            "title": title,
            "description": desc,
            "pr": {
                "url": prurl,
                "number": pr_num
            }
        })


    return stories

# ---------------------------
# Canonical assembly
# ---------------------------

def build_canonical(
    master_path: str,
    sql_root: str,
    business_paths: List[str],
    master_sheet: Optional[str] = None,
    highlights_max_rows: int = 5000,
    highlights_max_cols: int = 200,
) -> Dict[str, Any]:

    df = read_master_excel(master_path, sheet=master_sheet)
    stories = master_rows_to_stories(df)

    # Pre-extract all business highlights once
    business_inputs = []
    for bp in business_paths:
        if not os.path.isfile(bp):
            continue
        business_inputs.append(
            extract_highlights_from_workbook(
                bp,
                max_rows=highlights_max_rows,
                max_cols=highlights_max_cols
            )
        )

    # Attach SQL + relevant business highlights to each story
    canonical_stories = []
    for s in stories:
        story_id = s.get("story_id", "")
        pr_num = s.get("pr", {}).get("number", "")

        keys = []
        if story_id:
            keys.append(story_id)
            # also include digits only to match filenames like 12345678.sql
            keys.append(re.sub(r"\D", "", story_id))
        if pr_num:
            keys.append(pr_num)

        sql_files = find_sql_files(sql_root, keys)
        sql_payloads = []
        impacted_tables = set()
        operations = set()
        declared_objects = []

        for fp in sql_files:
            text = safe_read_text(fp)
            summary = parse_sql_summary(text)
            impacted_tables.update(summary.get("tables_referenced", []))
            operations.update(summary.get("operations", []))
            declared_objects.extend(summary.get("objects_declared", []))

            sql_payloads.append({
                "path": fp,
                "file_name": os.path.basename(fp),
                "sha256": sha256_file(fp),
                "summary": summary,
                # Keep full SQL text (if you want). If file sizes are huge, you can drop this field.
                "sql_text": text
            })

        # Business highlights can be associated by story id appearing anywhere in row snapshot values
        matched_business = []
        for wb in business_inputs:
            wb_name = wb.get("workbook", "").upper()

            # Match based on release (V20, V21, etc.)
            if release and release in wb_name:
                matched_business.append(wb)

        canonical_stories.append({
            **s,
            "sql": {
                "matched_files_count": len(sql_payloads),
                "files": sql_payloads,
                "derived": {
                    "tables_referenced": sorted(impacted_tables),
                    "operations": sorted(operations),
                    "objects_declared": declared_objects
                }
            },
            "business_inputs": {
                "workbooks_provided_count": len(business_inputs),
                "matched_workbooks_count": len(matched_business),
                "matched": matched_business,
                # Keep all highlights too (optional). Comment out if too large.
                "all": business_inputs
            }
        })

    return {
        "schema_version": "1.0",
        "generated_at": now_iso(),
        "sources": {
            "master_excel": master_path,
            "sql_root": sql_root,
            "business_excels": business_paths
        },
        "stories_count": len(canonical_stories),
        "stories": canonical_stories
    }

# ---------------------------
# CLI
# ---------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", required=True, help="Path to master Excel (.xlsx)")
    ap.add_argument("--master-sheet", default=None, help="Optional sheet name in master Excel")
    ap.add_argument("--sql-root", required=True, help="Folder containing downloaded SQL scripts")
    ap.add_argument("--business", action="append", default=[], help="Business Excel(s) with highlights; pass multiple --business")
    ap.add_argument("--out", required=True, help="Output canonical JSON file path")
    ap.add_argument("--highlights-max-rows", type=int, default=5000)
    ap.add_argument("--highlights-max-cols", type=int, default=200)
    args = ap.parse_args()

    canonical = build_canonical(
        master_path=args.master,
        sql_root=args.sql_root,
        business_paths=args.business or [],
        master_sheet=args.master_sheet,
        highlights_max_rows=args.highlights_max_rows,
        highlights_max_cols=args.highlights_max_cols,
    )

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(canonical, f, indent=2, ensure_ascii=False)

    print(f"✅ Wrote canonical JSON: {args.out}")
    print(f"Stories: {canonical['stories_count']}")
    # Print a small sanity summary
    if canonical["stories"]:
        print("Sample story keys:", list(canonical["stories"][0].keys()))

if __name__ == "__main__":
    main()

