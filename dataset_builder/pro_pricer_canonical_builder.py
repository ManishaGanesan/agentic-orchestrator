from __future__ import annotations

import json
import logging
import os
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# Optional SQL formatter
try:
    import sqlparse
except ImportError:
    sqlparse = None


# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------

# Common yellow fills seen in Excel workbooks.
YELLOW_RGBS = {
    "FFFFFF00",  # bright yellow
    "FFFF00",    # sometimes seen without alpha
    "FFFFFF99",  # pale yellow
    "FFFFEB9C",  # Excel light yellow
}

# Header aliases
KEY_STATE_COLUMNS = {"State", "State Id", "State_ID", "StateID"}
KEY_EFFECTIVE_DATE_COLUMNS = {"Effective Date", "EffDate", "Eff Date"}
KEY_PROC_DESC_COLUMNS = {"Procedure Description", "PDescription"}
KEY_PROC_VAR_NAME_COLUMNS = {
    "Procedure/ Variable Name",
    "Procedure/Variable Name",
    "Procedure Variable Name",
    "PCode",
}


# ------------------------------------------------------------------
# Data contracts
# ------------------------------------------------------------------

@dataclass
class ProcedureItem:
    display_order: Any
    pcode: Any
    source_column: str


@dataclass
class StateProcedureRecord:
    state_name: Optional[str]
    state_acronym: Optional[Any]
    state_id: Optional[Any]
    effective_date: Optional[Any]
    action: str  # ADD / UPDATE
    procedures: List[Dict[str, Any]]
    source_file: str


@dataclass
class ProcedureVariableItem:
    variable_name: str
    updated: bool


@dataclass
class ProcedureVariableRecord:
    action: str  # ADD / UPDATE
    pdescription: Dict[str, Any]
    pcode: Dict[str, Any]
    variables: List[Dict[str, Any]]
    source_file: str


@dataclass
class SqlScriptRecord:
    file_name: str
    content: str


@dataclass
class ValidationRecord:
    has_state_procedures_file: bool
    has_variables_file: bool
    has_sql_script_file: bool
    state_procedures_records_count: int
    procedure_variables_records_count: int
    sql_script_loaded: bool


@dataclass
class UpdateFolderRecord:
    update_folder: str
    status: str  # processed / skipped / failed
    state_procedures: List[Dict[str, Any]]
    procedure_variables: List[Dict[str, Any]]
    sql_script: Optional[Dict[str, Any]]
    validation: Dict[str, Any]
    errors: List[str]


# ------------------------------------------------------------------
# Processor
# ------------------------------------------------------------------

class ProPricerCanonicalBuilder:
    """
    Builds canonical JSON from a folder structure like:

      C:\\Users\\dparipat\\ProPricerSheets
        └── 2026
            └── V2605.01
                ├── MedicaidAPGPro_StateProcedure
                │   ├── *.xlsx / *.xlsm
                │   └── *.sql
                └── PhysicianPro_StateProcedure
                    ├── *.xlsx / *.xlsm
                    └── *.sql

    Supports update folders containing:
      - one StateProcedures workbook
      - one VariablesforProcedures workbook (optional)
      - one SQL script
    """

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    # --------------------------------------------------------------
    # Public API
    # --------------------------------------------------------------

    def process_root(self, root_path: str) -> Dict[str, Any]:
        if not os.path.isdir(root_path):
            raise FileNotFoundError(f"Root folder not found: {root_path}")

        result: Dict[str, Any] = {
            "years": {},
            "summary": {
                "years_count": 0,
                "versions_count": 0,
                "update_folders_count": 0,
                "processed_count": 0,
                "skipped_count": 0,
                "failed_count": 0,
            },
        }

        year_count = 0
        version_count = 0
        update_folder_count = 0
        processed_count = 0
        skipped_count = 0
        failed_count = 0

        for year_name in sorted(os.listdir(root_path)):
            year_path = os.path.join(root_path, year_name)

            if not os.path.isdir(year_path):
                continue

            # Only process YYYY folder names
            if not re.fullmatch(r"\d{4}", year_name):
                self.logger.debug("Skipping non-year folder: %s", year_path)
                continue

            self.logger.info("Processing year folder: %s", year_path)
            result["years"][year_name] = {}
            year_count += 1

            for version_name in sorted(os.listdir(year_path)):
                version_path = os.path.join(year_path, version_name)

                if not os.path.isdir(version_path):
                    continue

                # Version format like V2605.01
                if not re.fullmatch(r"V\d{4}\.\d{2}", version_name, flags=re.IGNORECASE):
                    self.logger.debug("Skipping non-version folder: %s", version_path)
                    continue

                self.logger.info("Processing version folder: %s", version_path)
                version_count += 1
                version_updates: List[Dict[str, Any]] = []

                for update_folder_name in sorted(os.listdir(version_path)):
                    update_folder_path = os.path.join(version_path, update_folder_name)

                    if not os.path.isdir(update_folder_path):
                        continue

                    update_folder_count += 1
                    update_record = self._safe_process_update_folder(update_folder_path)
                    version_updates.append(update_record)

                    status = update_record.get("status")
                    if status == "processed":
                        processed_count += 1
                    elif status == "skipped":
                        skipped_count += 1
                    elif status == "failed":
                        failed_count += 1

                result["years"][year_name][version_name] = version_updates

        result["summary"] = {
            "years_count": year_count,
            "versions_count": version_count,
            "update_folders_count": update_folder_count,
            "processed_count": processed_count,
            "skipped_count": skipped_count,
            "failed_count": failed_count,
        }

        return result

    # --------------------------------------------------------------
    # Safe wrapper
    # --------------------------------------------------------------

    def _safe_process_update_folder(self, folder_path: str) -> Dict[str, Any]:
        folder_name = os.path.basename(folder_path)

        try:
            return self._process_update_folder(folder_path)
        except Exception as ex:
            self.logger.exception("Failed processing folder: %s", folder_path)
            return asdict(
                UpdateFolderRecord(
                    update_folder=folder_name,
                    status="failed",
                    state_procedures=[],
                    procedure_variables=[],
                    sql_script=None,
                    validation=asdict(
                        ValidationRecord(
                            has_state_procedures_file=False,
                            has_variables_file=False,
                            has_sql_script_file=False,
                            state_procedures_records_count=0,
                            procedure_variables_records_count=0,
                            sql_script_loaded=False,
                        )
                    ),
                    errors=[str(ex)],
                )
            )

    # --------------------------------------------------------------
    # Folder-level processing
    # --------------------------------------------------------------

    def _process_update_folder(self, folder_path: str) -> Dict[str, Any]:
        self.logger.info("Processing update folder: %s", folder_path)

        files = [
            f for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f))
        ]

        state_proc_file: Optional[str] = None
        variables_file: Optional[str] = None
        sql_file: Optional[str] = None
        errors: List[str] = []

        for file_name in files:
            lower = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            # More flexible detection than strict suffix only
            is_excel = lower.endswith(".xlsx") or lower.endswith(".xlsm")

            if is_excel and "stateprocedures" in lower:
                if state_proc_file is None:
                    state_proc_file = full_path
                else:
                    errors.append(
                        f"Multiple StateProcedures files found. Using first: {os.path.basename(state_proc_file)}"
                    )

            elif is_excel and (
                "variablesforprocedures" in lower or "variablesforprocedres" in lower
            ):
                if variables_file is None:
                    variables_file = full_path
                else:
                    errors.append(
                        f"Multiple Variables files found. Using first: {os.path.basename(variables_file)}"
                    )

            elif lower.endswith(".sql"):
                if sql_file is None:
                    sql_file = full_path
                else:
                    errors.append(
                        f"Multiple SQL files found. Using first: {os.path.basename(sql_file)}"
                    )

        has_state_procedures_file = state_proc_file is not None
        has_variables_file = variables_file is not None
        has_sql_script_file = sql_file is not None

        state_procedures: List[Dict[str, Any]] = []
        procedure_variables: List[Dict[str, Any]] = []
        sql_script: Optional[Dict[str, Any]] = None

        if state_proc_file:
            try:
                self.logger.info("Reading StateProcedures workbook: %s", state_proc_file)
                state_procedures = self._process_state_procedures(state_proc_file)
            except Exception as ex:
                msg = f"Error processing StateProcedures file {os.path.basename(state_proc_file)}: {str(ex)}"
                self.logger.exception(msg)
                errors.append(msg)

        if variables_file:
            try:
                self.logger.info("Reading Variables workbook: %s", variables_file)
                procedure_variables = self._process_variables(variables_file)
            except Exception as ex:
                msg = f"Error processing Variables file {os.path.basename(variables_file)}: {str(ex)}"
                self.logger.exception(msg)
                errors.append(msg)

        if sql_file:
            try:
                self.logger.info("Reading SQL file: %s", sql_file)
                sql_script = self._read_sql_script(sql_file)
            except Exception as ex:
                msg = f"Error reading SQL file {os.path.basename(sql_file)}: {str(ex)}"
                self.logger.exception(msg)
                errors.append(msg)

        validation = asdict(
            ValidationRecord(
                has_state_procedures_file=has_state_procedures_file,
                has_variables_file=has_variables_file,
                has_sql_script_file=has_sql_script_file,
                state_procedures_records_count=len(state_procedures),
                procedure_variables_records_count=len(procedure_variables),
                sql_script_loaded=sql_script is not None,
            )
        )

        if not has_state_procedures_file and not has_variables_file and not has_sql_script_file:
            status = "skipped"
        elif errors and not state_procedures and not procedure_variables and not sql_script:
            status = "failed"
        else:
            status = "processed"

        return asdict(
            UpdateFolderRecord(
                update_folder=os.path.basename(folder_path),
                status=status,
                state_procedures=state_procedures,
                procedure_variables=procedure_variables,
                sql_script=sql_script,
                validation=validation,
                errors=errors,
            )
        )

    def _read_sql_script(self, sql_file_path: str) -> Dict[str, Any]:
        raw_content = ""

        #1. Read the file content first
        try:
            with open(sql_file_path, "r", encoding="utf-8") as f:
                raw_content = f.read()
        except UnicodeDecodeError:
            with open(sql_file_path, "r", encoding="cp1252", errors="replace") as f:
                raw_content = f.read()

        #2. Process and format the statements line-by-line
        try:

            statements = raw_content.splitlines()
            formatted_lines = []

            for stmt in statements:
                if stmt.strip(): #skip completely empty lines
                    #2. Format the individual statement
                    f_stmt = sqlparse.format(
                        stmt,
                        reindent=True,
		                reindent_aligned=True,
                        indent_width=4,
                        keyword_case="upper",
                        strip_comments=False,
                        use_space_around_operators=True
                    )
                    # remove any trailing newlines sqlparse adds to individual statements
                    formatted_lines.append(f_stmt.rstrip())

        #3 Join them all back together with a clean line break
            formatted_content = "\n".join(formatted_lines)
        except Exception as ex:
                self.logger.warning(
                    "SQL formatting failed for %s. Using raw SQL. Error: %s",
                    sql_file_path,
                    str(ex),
                )

        return asdict(
            SqlScriptRecord(
                file_name=os.path.basename(sql_file_path),
                content=formatted_content,
            )
        )

    # --------------------------------------------------------------
    # State Procedures logic
    # --------------------------------------------------------------

    def _process_state_procedures(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = self._read_headers(ws)
        state_name = self._extract_state_name_from_title(file_path)

        output: List[Dict[str, Any]] = []

        for row_idx in range(2, ws.max_row + 1):
            row_map = self._row_to_dict(ws, headers, row_idx)
            highlighted_columns = self._get_highlighted_columns(ws, headers, row_idx)

            if not highlighted_columns:
                continue

            action = "ADD" if self._is_full_data_row_highlighted(ws, headers, row_idx) else "UPDATE"

            state_id = self._get_first_value(row_map, KEY_STATE_COLUMNS)
            effective_date = self._get_first_value(row_map, KEY_EFFECTIVE_DATE_COLUMNS)

            # Acronym only from highlighted state column
            state_acronym = None
            normalized_state_headers = {self._normalize(h) for h in KEY_STATE_COLUMNS}
            for col_name in highlighted_columns:
                if self._normalize(col_name) in normalized_state_headers:
                    state_acronym = row_map.get(col_name)
                    break

            procedures: List[Dict[str, Any]] = []

            for col_name in highlighted_columns:
                if self._is_procedure_column(col_name):
                    pcode = row_map.get(col_name)

                    if self._has_meaningful_value(pcode):
                        procedures.append(
                            asdict(
                                ProcedureItem(
                                    display_order=self._extract_display_order(col_name),
                                    pcode=pcode,
                                    source_column=col_name.strip(),
                                )
                            )
                        )

            # Keep only rows that actually changed procedure columns
            if not procedures:
                continue

            output.append(
                asdict(
                    StateProcedureRecord(
                        state_name=state_name,
                        state_acronym=state_acronym,
                        state_id=state_id,
                        effective_date=effective_date,
                        action=action,
                        procedures=procedures,
                        source_file=os.path.basename(file_path),
                    )
                )
            )

        return output

    # --------------------------------------------------------------
    # Variables logic
    # --------------------------------------------------------------

    def _process_variables(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = self._read_headers(ws)
        output: List[Dict[str, Any]] = []

        pdesc_col = self._find_first_header(headers, KEY_PROC_DESC_COLUMNS)
        pcode_col = self._find_first_header(headers, KEY_PROC_VAR_NAME_COLUMNS)

        for row_idx in range(2, ws.max_row + 1):
            row_map = self._row_to_dict(ws, headers, row_idx)
            highlighted_columns = self._get_highlighted_columns(ws, headers, row_idx)

            if not highlighted_columns:
                continue

            action = "ADD" if self._is_full_data_row_highlighted(ws, headers, row_idx) else "UPDATE"

            pdesc_value = row_map.get(pdesc_col) if pdesc_col else None
            pcode_value = row_map.get(pcode_col) if pcode_col else None

            variables: List[Dict[str, Any]] = []

            # X logic is separate from highlight logic
            for col_name in headers:
                if pdesc_col and self._normalize(col_name) == self._normalize(pdesc_col):
                    continue
                if pcode_col and self._normalize(col_name) == self._normalize(pcode_col):
                    continue

                cell_value = row_map.get(col_name)

                if str(cell_value).strip().upper() == "X":
                    variables.append(
                        asdict(
                            ProcedureVariableItem(
                                variable_name=col_name.strip(),
                                updated=(col_name in highlighted_columns),
                            )
                        )
                    )

            output.append(
                asdict(
                    ProcedureVariableRecord(
                        action=action,
                        pdescription={
                            "value": pdesc_value,
                            "updated": (pdesc_col in highlighted_columns) if pdesc_col else False,
                        },
                        pcode={
                            "value": pcode_value,
                            "updated": (pcode_col in highlighted_columns) if pcode_col else False,
                        },
                        variables=variables,
                        source_file=os.path.basename(file_path),
                    )
                )
            )

        return output

    # --------------------------------------------------------------
    # Excel helpers
    # --------------------------------------------------------------

    def _read_headers(self, ws: Worksheet) -> List[str]:
        """
        Read headers from row 1 and trim spaces.
        """
        headers: List[str] = []

        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value is None:
                header = f"COL_{col_idx}"
            else:
                header = str(value).strip()

            headers.append(header)

        return headers

    def _row_to_dict(self, ws: Worksheet, headers: List[str], row_idx: int) -> Dict[str, Any]:
        """
        Build row map using already-trimmed headers.
        """
        row_data: Dict[str, Any] = {}

        for col_idx, header in enumerate(headers, start=1):
            row_data[header.strip()] = ws.cell(row=row_idx, column=col_idx).value

        return row_data

    def _get_highlighted_columns(self, ws: Worksheet, headers: List[str], row_idx: int) -> List[str]:
        """
        Return header names for yellow-highlighted cells in the row.
        """
        highlighted: List[str] = []

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if self._is_yellow_fill(cell):
                highlighted.append(header.strip())

        return highlighted

    def _is_full_data_row_highlighted(self, ws: Worksheet, headers: List[str], row_idx: int) -> bool:
        """
        Full row highlighted => all non-empty cells in the row are yellow.
        """
        non_empty_cells = 0
        highlighted_non_empty_cells = 0

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if self._has_meaningful_value(cell.value):
                non_empty_cells += 1
                if self._is_yellow_fill(cell):
                    highlighted_non_empty_cells += 1

        return non_empty_cells > 0 and non_empty_cells == highlighted_non_empty_cells

    def _is_yellow_fill(self, cell: Cell) -> bool:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return False

        fg = getattr(fill, "fgColor", None)
        if fg:
            rgb = getattr(fg, "rgb", None)
            if rgb and str(rgb).upper() in YELLOW_RGBS:
                return True

            indexed = getattr(fg, "indexed", None)
            if indexed in {5, 6}:  # common yellow-ish indexed values
                return True

        start = getattr(fill, "start_color", None)
        if start:
            rgb = getattr(start, "rgb", None)
            if rgb and str(rgb).upper() in YELLOW_RGBS:
                return True

            indexed = getattr(start, "indexed", None)
            if indexed in {5, 6}:
                return True

        return False

    # --------------------------------------------------------------
    # Business helpers
    # --------------------------------------------------------------

    def _normalize(self, text: str) -> str:
        """
        Trim + normalize spaces + lowercase for all comparisons.
        """
        if text is None:
            return ""
        return re.sub(r"\s+", " ", str(text)).strip().lower()

    def _extract_state_name_from_title(self, file_path: str) -> str:
        """
        Best-effort extraction from filename.

        Example:
          V2404.00 - New State Procedures addons and arrays for Florida APR-DRG (...)_StateProcedures.xlsx
        """
        stem = Path(file_path).stem

        cleaned = re.sub(r"_StateProcedures$", "", stem, flags=re.IGNORECASE)
        cleaned = re.sub(r"_VariablesforProcedures$", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"_VariablesforProcedres$", "", cleaned, flags=re.IGNORECASE)

        match = re.search(
            r"\bfor\s+([A-Za-z][A-Za-z\s-]+?)(?:\s+APR|\s+DRG|\s+\(|$)",
            cleaned,
            flags=re.IGNORECASE,
        )
        if match:
            return match.group(1).strip()

        return cleaned.strip()

    def _get_first_value(self, row_map: Dict[str, Any], candidate_headers: set[str]) -> Any:
        """
        Compare by normalized keys so aliases match even with extra spaces.
        """
        normalized_targets = {self._normalize(h) for h in candidate_headers}

        for key in row_map:
            if self._normalize(key) in normalized_targets:
                return row_map.get(key)

        return None

    def _find_first_header(self, headers: List[str], candidate_headers: set[str]) -> Optional[str]:
        """
        Compare by normalized headers so aliases match even with extra spaces.
        """
        normalized_targets = {self._normalize(h) for h in candidate_headers}

        for header in headers:
            if self._normalize(header) in normalized_targets:
                return header

        return None

    def _is_procedure_column(self, column_name: str) -> bool:
        """
        Procedure columns are expected to start with 'Procedure'
        but should NOT be:
          - Procedure Description
          - Procedure/Variable Name
          - Procedure Variable Name
        """
        col = self._normalize(column_name)

        excluded = {
            self._normalize("Procedure Description"),
            self._normalize("Procedure/Variable Name"),
            self._normalize("Procedure/ Variable Name"),
            self._normalize("Procedure Variable Name"),
        }

        return col.startswith("procedure") and col not in excluded

    def _extract_display_order(self, column_name: str) -> Any:
        """
        Examples:
          Procedure 1   -> 1
          Procedure_02  -> 2
          otherwise     -> original column name
        """
        match = re.search(r"(\d+)", str(column_name))
        return int(match.group(1)) if match else column_name.strip()

    def _has_meaningful_value(self, value: Any) -> bool:
        if value is None:
            return False

        if isinstance(value, str) and value.strip() == "":
            return False

        return True


# ------------------------------------------------------------------
# Logging setup
# ------------------------------------------------------------------

def setup_logger(log_file_path: str) -> logging.Logger:
    logger = logging.getLogger("pro_pricer_canonical_builder")
    logger.setLevel(logging.INFO)

    if logger.handlers:
        return logger

    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(log_file_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

if __name__ == "__main__":
    ROOT_FOLDER = r"../input_excels/ProPricerSheets"
    OUTPUT_FILE = r"./out/canonical_output.json"
    LOG_FILE = r"./out/canonical_output.log"

    logger = setup_logger(LOG_FILE)
    logger.info("Starting canonical JSON build...")

    builder = ProPricerCanonicalBuilder(logger)

    try:
        result = builder.process_root(ROOT_FOLDER)

        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=4, default=str)

        logger.info("Canonical JSON written to: %s", OUTPUT_FILE)
        logger.info("Summary: %s", json.dumps(result.get("summary", {}), indent=2))

        print(f"✅ Canonical JSON written to: {OUTPUT_FILE}")
        print(f"✅ Log file written to: {LOG_FILE}")

    except Exception as ex:
        logger.exception("Fatal error while building canonical JSON.")
        print(f"❌ Fatal error: {str(ex)}")