from __future__ import annotations

import os
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------

# Common yellow fills seen in Excel workbooks.
# If your workbook uses a different shade, print the cell color and add it here.
YELLOW_RGBS = {
    "FFFFFF00",  # bright yellow
    "FFFF00",    # sometimes seen without alpha
    "FFFFFF99",  # pale yellow
    "FFFFEB9C",  # Excel light yellow
}

# Header aliases
KEY_STATE_COLUMNS = {"State", "State Id", "State_ID", "StateID"}
KEY_EFFECTIVE_DATE_COLUMNS = {"Effective Date", "EffDate", "Eff Date"}
KEY_PROC_DESC_COLUMNS = {"Procedure Description", "PDescription"}
KEY_PROC_VAR_NAME_COLUMNS = {
    "Procedure/ Variable Name",
    "Procedure/Variable Name",
    "Procedure Variable Name",
    "PCode",
}


# ------------------------------------------------------------------
# Data contracts
# ------------------------------------------------------------------

@dataclass
class ProcedureItem:
    display_order: Any
    pcode: Any
    source_column: str


@dataclass
class StateProcedureRecord:
    state_name: Optional[str]
    state_acronym: Optional[Any]
    state_id: Optional[Any]
    effective_date: Optional[Any]
    action: str  # ADD / UPDATE
    procedures: List[Dict[str, Any]]
    source_file: str


@dataclass
class ProcedureVariableItem:
    variable_name: str
    updated: bool


@dataclass
class ProcedureVariableRecord:
    action: str  # ADD / UPDATE
    pdescription: Dict[str, Any]
    pcode: Dict[str, Any]
    variables: List[Dict[str, Any]]
    source_file: str


# ------------------------------------------------------------------
# Processor
# ------------------------------------------------------------------

class InputProcessor:
    """
    Handles Excel input and converts it into structured JSON
    for downstream agents.

    Supported files inside input folder:
      - *_StateProcedures.xlsx / .xlsm
      - *_VariablesforProcedures.xlsx / .xlsm

    Also tolerates the older typo:
      - *_VariablesforProcedres.xlsx / .xlsm
    """

    # --------------------------------------------------------------
    # Public APIs
    # --------------------------------------------------------------

    def process(
        self,
        state_proc_file: Optional[str] = None,
        variables_file: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Process explicitly provided files.
        Either file can be None.
        """
        result = {
            "state_procedures": [],
            "procedure_variables": []
        }

        if state_proc_file:
            result["state_procedures"].extend(
                self._process_state_procedures(state_proc_file)
            )

        if variables_file:
            result["procedure_variables"].extend(
                self._process_variables(variables_file)
            )

        return result

    def process_folder(self, folder_path: str) -> Dict[str, Any]:
        """
        Process all supported Excel files found in a folder.

        Rules:
        - files ending with _StateProcedures.xlsx/.xlsm => state procedures logic
        - files ending with _VariablesforProcedures.xlsx/.xlsm => variables logic
        - files ending with _VariablesforProcedres.xlsx/.xlsm => variables logic (legacy typo)
        - any other file => ignored
        """
        result = {
            "state_procedures": [],
            "procedure_variables": []
        }

        if not os.path.isdir(folder_path):
            raise FileNotFoundError(f"Folder not found: {folder_path}")

        files = [
            f for f in os.listdir(folder_path)
            if f.lower().endswith((".xlsx", ".xlsm"))
        ]

        for file_name in files:
            file_lower = file_name.lower()
            full_path = os.path.join(folder_path, file_name)

            if file_lower.endswith("_stateprocedures.xlsx") or file_lower.endswith("_stateprocedures.xlsm"):
                print(f"Processing State Procedures file: {file_name}")
                result["state_procedures"].extend(
                    self._process_state_procedures(full_path)
                )

            elif (
                file_lower.endswith("_variablesforprocedures.xlsx")
                or file_lower.endswith("_variablesforprocedures.xlsm")
                or file_lower.endswith("_variablesforprocedres.xlsx")
                or file_lower.endswith("_variablesforprocedres.xlsm")
            ):
                print(f"Processing Variables file: {file_name}")
                result["procedure_variables"].extend(
                    self._process_variables(full_path)
                )

            else:
                print(f"Skipping unsupported file: {file_name}")

        return result

    # --------------------------------------------------------------
    # State Procedures logic
    # --------------------------------------------------------------

    def _process_state_procedures(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = self._read_headers(ws)
        state_name = self._extract_state_name_from_title(file_path)

        output: List[Dict[str, Any]] = []

        for row_idx in range(2, ws.max_row + 1):
            row_map = self._row_to_dict(ws, headers, row_idx)
            highlighted_columns = self._get_highlighted_columns(ws, headers, row_idx)

            if not highlighted_columns:
                continue

            action = "ADD" if self._is_full_data_row_highlighted(ws, headers, row_idx) else "UPDATE"

            state_id = self._get_first_value(row_map, KEY_STATE_COLUMNS)
            effective_date = self._get_first_value(row_map, KEY_EFFECTIVE_DATE_COLUMNS)

            # Acronym: fetch from highlighted state column (as per note)
            state_acronym = None
            normalized_state_headers = {self._normalize(h) for h in KEY_STATE_COLUMNS}
            for col_name in highlighted_columns:
                if self._normalize(col_name) in normalized_state_headers:
                    state_acronym = row_map.get(col_name)
                    break

            procedures: List[Dict[str, Any]] = []

            for col_name in highlighted_columns:
                if self._is_procedure_column(col_name):
                    pcode = row_map.get(col_name)

                    if self._has_meaningful_value(pcode):
                        procedures.append(
                            asdict(
                                ProcedureItem(
                                    display_order=self._extract_display_order(col_name),
                                    pcode=pcode,
                                    source_column=col_name.strip()
                                )
                            )
                        )

            # Keep only rows that actually affect procedure columns
            if not procedures:
                continue

            output.append(
                asdict(
                    StateProcedureRecord(
                        state_name=state_name,
                        state_acronym=state_acronym,
                        state_id=state_id,
                        effective_date=effective_date,
                        action=action,
                        procedures=procedures,
                        source_file=os.path.basename(file_path)
                    )
                )
            )

        return output

    # --------------------------------------------------------------
    # Variables logic
    # --------------------------------------------------------------

    def _process_variables(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        headers = self._read_headers(ws)
        output: List[Dict[str, Any]] = []

        pdesc_col = self._find_first_header(headers, KEY_PROC_DESC_COLUMNS)
        pcode_col = self._find_first_header(headers, KEY_PROC_VAR_NAME_COLUMNS)

        for row_idx in range(2, ws.max_row + 1):
            row_map = self._row_to_dict(ws, headers, row_idx)
            highlighted_columns = self._get_highlighted_columns(ws, headers, row_idx)

            if not highlighted_columns:
                continue

            action = "ADD" if self._is_full_data_row_highlighted(ws, headers, row_idx) else "UPDATE"

            pdesc_value = row_map.get(pdesc_col) if pdesc_col else None
            pcode_value = row_map.get(pcode_col) if pcode_col else None

            variables: List[Dict[str, Any]] = []

            # X logic is separate from highlight logic
            for col_name in headers:
                if pdesc_col and self._normalize(col_name) == self._normalize(pdesc_col):
                    continue
                if pcode_col and self._normalize(col_name) == self._normalize(pcode_col):
                    continue

                cell_value = row_map.get(col_name)

                if str(cell_value).strip().upper() == "X":
                    variables.append(
                        asdict(
                            ProcedureVariableItem(
                                variable_name=col_name.strip(),
                                updated=(col_name in highlighted_columns)
                            )
                        )
                    )

            output.append(
                asdict(
                    ProcedureVariableRecord(
                        action=action,
                        pdescription={
                            "value": pdesc_value,
                            "updated": (pdesc_col in highlighted_columns) if pdesc_col else False
                        },
                        pcode={
                            "value": pcode_value,
                            "updated": (pcode_col in highlighted_columns) if pcode_col else False
                        },
                        variables=variables,
                        source_file=os.path.basename(file_path)
                    )
                )
            )

        return output

    # --------------------------------------------------------------
    # Excel helpers
    # --------------------------------------------------------------

    def _read_headers(self, ws: Worksheet) -> List[str]:
        """
        Read headers from row 1.
        Trim spaces at read time so Excel-side headers are cleaned early.
        """
        headers: List[str] = []

        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value is None:
                header = f"COL_{col_idx}"
            else:
                header = str(value).strip()

            headers.append(header)

        return headers

    def _row_to_dict(self, ws: Worksheet, headers: List[str], row_idx: int) -> Dict[str, Any]:
        """
        Build row map using already trimmed headers.
        """
        row_data: Dict[str, Any] = {}

        for col_idx, header in enumerate(headers, start=1):
            row_data[header.strip()] = ws.cell(row=row_idx, column=col_idx).value

        return row_data

    def _get_highlighted_columns(self, ws: Worksheet, headers: List[str], row_idx: int) -> List[str]:
        """
        Return trimmed header names for highlighted cells in the row.
        """
        highlighted: List[str] = []

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if self._is_yellow_fill(cell):
                highlighted.append(header.strip())

        return highlighted

    def _is_full_data_row_highlighted(self, ws: Worksheet, headers: List[str], row_idx: int) -> bool:
        """
        Full row highlighted => all non-empty cells in the row are yellow.
        """
        non_empty_cells = 0
        highlighted_non_empty_cells = 0

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if self._has_meaningful_value(cell.value):
                non_empty_cells += 1
                if self._is_yellow_fill(cell):
                    highlighted_non_empty_cells += 1

        return non_empty_cells > 0 and non_empty_cells == highlighted_non_empty_cells

    def _is_yellow_fill(self, cell: Cell) -> bool:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return False

        fg = getattr(fill, "fgColor", None)
        if fg:
            rgb = getattr(fg, "rgb", None)
            if rgb and str(rgb).upper() in YELLOW_RGBS:
                return True

            indexed = getattr(fg, "indexed", None)
            if indexed in {5, 6}:  # common yellow-ish indexed values
                return True

        start = getattr(fill, "start_color", None)
        if start:
            rgb = getattr(start, "rgb", None)
            if rgb and str(rgb).upper() in YELLOW_RGBS:
                return True

            indexed = getattr(start, "indexed", None)
            if indexed in {5, 6}:
                return True

        return False

    # --------------------------------------------------------------
    # Business helpers
    # --------------------------------------------------------------

    def _normalize(self, text: str) -> str:
        """
        Trim + normalize spaces + lowercase for all comparisons.
        This ensures Excel headers and alias keys match even if spacing varies.
        """
        if text is None:
            return ""
        return re.sub(r"\s+", " ", str(text)).strip().lower()

    def _extract_state_name_from_title(self, file_path: str) -> str:
        """
        Best-effort extraction from filename.

        Example:
          V2404.00 - New State Procedures addons and arrays for Florida APR-DRG (...)_StateProcedures.xlsx

        Tries to extract 'Florida' from 'for Florida ...'
        Falls back to cleaned title without suffix.
        """
        stem = Path(file_path).stem

        cleaned = re.sub(r"_StateProcedures$", "", stem, flags=re.IGNORECASE)
        cleaned = re.sub(r"_VariablesforProcedures$", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"_VariablesforProcedres$", "", cleaned, flags=re.IGNORECASE)

        match = re.search(
            r"\bfor\s+([A-Za-z][A-Za-z\s-]+?)(?:\s+APR|\s+DRG|\s+\(|$)",
            cleaned,
            flags=re.IGNORECASE
        )
        if match:
            return match.group(1).strip()

        return cleaned.strip()

    def _get_first_value(self, row_map: Dict[str, Any], candidate_headers: set[str]) -> Any:
        """
        Compare by normalized keys so aliases match even with extra spaces.
        """
        normalized_targets = {self._normalize(h) for h in candidate_headers}

        for key in row_map:
            if self._normalize(key) in normalized_targets:
                return row_map.get(key)

        return None

    def _find_first_header(self, headers: List[str], candidate_headers: set[str]) -> Optional[str]:
        """
        Compare by normalized headers so aliases match even with extra spaces.
        """
        normalized_targets = {self._normalize(h) for h in candidate_headers}

        for header in headers:
            if self._normalize(header) in normalized_targets:
                return header

        return None

    def _is_procedure_column(self, column_name: str) -> bool:
        """
        Procedure columns are expected to start with 'Procedure'
        but should NOT be:
          - Procedure Description
          - Procedure/Variable Name
          - Procedure Variable Name

        Comparison is done using normalized trimmed values.
        """
        col = self._normalize(column_name)

        excluded = {
            self._normalize("Procedure Description"),
            self._normalize("Procedure/Variable Name"),
            self._normalize("Procedure/ Variable Name"),
            self._normalize("Procedure Variable Name"),
        }

        return col.startswith("procedure") and col not in excluded

    def _extract_display_order(self, column_name: str) -> Any:
        """
        Examples:
          Procedure 1   -> 1
          Procedure_02  -> 2
          otherwise     -> original column name
        """
        match = re.search(r"(\d+)", str(column_name))
        return int(match.group(1)) if match else column_name.strip()

    def _has_meaningful_value(self, value: Any) -> bool:
        if value is None:
            return False

        if isinstance(value, str) and value.strip() == "":
            return False

        return True